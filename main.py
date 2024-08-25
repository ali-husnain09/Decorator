from pyfiglet import Figlet, figlet_format
from colorama import Fore, Back, Style, init
import openpyxl
import time

# Initialize colorama
init(autoreset=True)


# ASCII art text
decorator_text = figlet_format("DECORATOR")
str1 = figlet_format("TRUE CODERS")
# Print with colors using Colorama
print(Fore.LIGHTRED_EX + decorator_text + Style.RESET_ALL, end="")
print(Fore.LIGHTGREEN_EX + "by " + Style.RESET_ALL)
print(Fore.LIGHTGREEN_EX + str1 + Style.RESET_ALL)
print(
    Back.LIGHTWHITE_EX
    + Fore.BLACK
    + "------------------------------------------COPYRIGHT©: ALI HUSNAIN "
    + Style.RESET_ALL
)
print(
    Fore.LIGHTWHITE_EX
    + "================================================================="
    + Style.RESET_ALL
)


# //here is to extract companies from sheet
class NameValidations:

    def __init__(self, details_path1, companies_path1, companies_list):
        self.details_path1 = details_path1
        self.companies_path1 = companies_path1
        self.companies_list = companies_list
        self.last_row_number = 0
        self.r_num = 0  # Initialize r_num as a class attribute

    def __checkValid__(self, name):
        name = str(name)
        valid_name = False
        valid_chars = []
        with open(self.companies_list, "r") as f:
            for line in f:
                valid_chars.append(line.strip())
        for index in valid_chars:
            if index in name.lower().split(" "):
                valid_name = True
                return valid_name
        return False

    def get_next_row_data(self):
        wb = openpyxl.load_workbook(self.details_path1)
        sheet = wb.active

        user_name = sheet.cell(row=self.last_row_number + 2, column=1).value
        property_address = sheet.cell(row=self.last_row_number + 2, column=2).value
        property_city = sheet.cell(row=self.last_row_number + 2, column=3).value
        property_state = sheet.cell(row=self.last_row_number + 2, column=4).value
        formal_address = sheet.cell(row=self.last_row_number + 2, column=5).value
        city = sheet.cell(row=self.last_row_number + 2, column=6).value
        state = sheet.cell(row=self.last_row_number + 2, column=7).value

        return (
            user_name,
            formal_address,
            city,
            state,
            property_address,
            property_city,
            property_state,
        )

    def save_value(self, sheet, value):
        sheet.cell(row=self.last_row_number + 2, column=8).value = value
        sheet.cell(row=self.last_row_number + 2, column=12).value = (
            self.last_row_number + 2
        )

    def appending_sheet(self, name, address, city, state, p_a, p_c, p_s):
        wb = openpyxl.load_workbook(self.companies_path1)
        sheet = wb.active
        sheet.cell(row=self.r_num + 2, column=1).value = name
        sheet.cell(row=self.r_num + 2, column=2).value = p_a
        sheet.cell(row=self.r_num + 2, column=3).value = p_c
        sheet.cell(row=self.r_num + 2, column=4).value = p_s
        sheet.cell(row=self.r_num + 2, column=5).value = address
        sheet.cell(row=self.r_num + 2, column=6).value = city
        sheet.cell(row=self.r_num + 2, column=7).value = state
        sheet.cell(row=self.r_num + 2, column=12).value = self.last_row_number + 2
        wb.save(self.companies_path1)
        self.r_num += 1  # Increment r_num after appending

    def company_checker(self):
        wb = openpyxl.load_workbook(self.details_path1)
        sheet = wb.active

        while True:
            user_name, formal_address, city, state, p_a, p_c, p_s = (
                self.get_next_row_data()
            )
            if not user_name:
                print(
                    Fore.BLACK
                    + Back.GREEN
                    + "\nNO MORE ROWS TO PROCESS"
                    + Style.RESET_ALL
                )
                print(
                    f"TOTAL NUMBER OF ROWS TO PROCESS: {Fore.LIGHTCYAN_EX}{self.last_row_number}{Style.RESET_ALL}"
                )
                break

            if self.__checkValid__(user_name):
                print(
                    f"{Fore.LIGHTBLUE_EX}COMPANY FOUND:{Style.RESET_ALL} {user_name}{Style.RESET_ALL}"
                )
                self.save_value(sheet, "It's a Company")
                self.appending_sheet(
                    user_name, formal_address, city, state, p_a, p_c, p_s
                )

            self.last_row_number += 1

        wb.save(self.details_path1)


# //here is the details merger
class Details_Merger:

    def __init__(self, details_path2, companies_path2):
        self.details_path2 = details_path2
        self.companies_path2 = companies_path2
        self.last_row_number = 0
        self.r_num = 0  # Initialize r_num as a class attribute

    def companies_sheet(self):
        wb = openpyxl.load_workbook(self.companies_path2)
        sheet = wb.active
        user_name = sheet.cell(row=self.last_row_number + 2, column=1).value
        phone_numbers = sheet.cell(row=self.last_row_number + 2, column=8).value
        emails = sheet.cell(row=self.last_row_number + 2, column=9).value
        agent_name = sheet.cell(row=self.last_row_number + 2, column=10).value
        agent_address = sheet.cell(row=self.last_row_number + 2, column=11).value
        to_row = sheet.cell(row=self.last_row_number + 2, column=12).value

        return user_name, phone_numbers, emails, agent_name, agent_address, to_row

    def details_sheet(self, ph, emails, ag_N, ag_A, to_row):
        wb = openpyxl.load_workbook(self.details_path2)
        sheet = wb.active
        row_number = int(to_row)
        sheet.cell(row=row_number, column=8).value = ph
        sheet.cell(row=row_number, column=9).value = emails
        sheet.cell(row=row_number, column=10).value = ag_N
        sheet.cell(row=row_number, column=11).value = ag_A
        sheet.cell(row=row_number, column=12).value = row_number
        print(f"THE PROVIDED VALUE SAVED IN: {row_number}")
        wb.save(self.details_path2)

    def merger(self):
        self.enteringDetails = True
        while self.enteringDetails:
            user, phone_numbers, emails, agent_name, agent_address, to_row = (
                self.companies_sheet()
            )
            if not user:
                print(
                    Fore.BLACK
                    + Back.GREEN
                    + "\nNO MORE ROWS TO PROCESS"
                    + Style.RESET_ALL
                )
                print(
                    f"TOTAL NUMBER OF ROWS TO PROCESS: {Fore.LIGHTCYAN_EX}{self.last_row_number}{Style.RESET_ALL}"
                )
                break
            self.details_sheet(phone_numbers, emails, agent_name, agent_address, to_row)

            self.last_row_number += 1


# //here is the eligibles merger


class Dectorator:

    def __init__(
        self,
        details_path1,
        companies_path1,
        details_path2,
        companies_path2,
        companies_list,
    ):
        self.details_sheet1 = details_path1
        self.details_sheet2 = details_path2
        self.companies_sheet1 = companies_path1
        self.companies_sheet2 = companies_path2
        self.companies_list = companies_list

    def decorating(self):
        self.enteringDetails = True
        while self.enteringDetails:
            print(
                f"\n{Fore.LIGHTYELLOW_EX}PRESS{Fore.LIGHTWHITE_EX} [1]: {Fore.LIGHTYELLOW_EX}TO EXTRACT THE COMPANIES{Style.RESET_ALL}"
            )
            print(
                f"{Fore.LIGHTYELLOW_EX}PRESS {Fore.LIGHTWHITE_EX}[2]: {Fore.LIGHTYELLOW_EX}TO MERGE THE DETAILS{Style.RESET_ALL}"
            )
            print(
                f"{Fore.LIGHTYELLOW_EX}PRESS {Fore.LIGHTWHITE_EX}[0]: {Fore.LIGHTYELLOW_EX}TO EXIT THE PROGRAM{Style.RESET_ALL}\n"
            )
            choice = input(f"{Fore.LIGHTGREEN_EX}ENTER YOUR CHOICE: {Style.RESET_ALL}")
            match choice:
                case "1":
                    start_time = time.time()
                    extractor_obj = NameValidations(
                        self.details_sheet1, self.companies_sheet1, self.companies_list
                    )
                    print(
                        f"\n==============================={Fore.LIGHTYELLOW_EX} [+] Start Processing{Style.RESET_ALL} ==============================="
                    )
                    print()
                    extractor_obj.company_checker()
                    end_time = time.time()
                    execution_time = end_time - start_time

                    print(
                        f"\nExecution time:{Fore.LIGHTYELLOW_EX} {execution_time} seconds {Style.RESET_ALL}"
                    )
                    print(
                        f"\n{Fore.LIGHTWHITE_EX}====================================================================================="
                    )
                case "2":
                    start_time = time.time()
                    merger_obj = Details_Merger(
                        self.details_sheet2, self.companies_sheet2
                    )
                    print(
                        f"\n==============================={Fore.LIGHTYELLOW_EX} [+] Start Processing{Style.RESET_ALL} ==============================="
                    )
                    print()
                    merger_obj.merger()
                    end_time = time.time()
                    execution_time = end_time - start_time

                    print(
                        f"\nExecution time:{Fore.LIGHTYELLOW_EX} {execution_time} seconds {Style.RESET_ALL}"
                    )
                    print(
                        f"\n{Fore.LIGHTWHITE_EX}====================================================================================="
                    )

                case "0":
                    self.enteringDetails = False
                    print(
                        f"\n{Fore.BLACK+ Back.LIGHTRED_EX} EXTING THE PROGRAM {Style.RESET_ALL}"
                    )
                case _:
                    print(
                        f"\n{Fore.BLACK+ Back.LIGHTRED_EX} INVALID CHOICE {Style.RESET_ALL}"
                    )


details_path1 = "Extractor/details.xlsx"
companies_path1 = "Extractor/companies.xlsx"
companies_list = "Extractor/companies.txt"

details_path2 = "Merger/details.xlsx"
companies_path2 = "Merger/companies.xlsx"

if __name__ == "__main__":
    decorator_obj = Dectorator(
        details_path1,
        companies_path1,
        details_path2,
        companies_path2,
        companies_list,
    )
    decorator_obj.decorating()
