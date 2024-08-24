import openpyxl
import colorama
from colorama import Fore, Back

colorama.init(autoreset=True)


class Eligibles_Merger:

    def __init__(self, companies_path, eligible_path):
        self.companies_path = companies_path
        self.eligible_path = eligible_path
        self.last_row_number = 0

    def eligible_sheet(self):
        wb = openpyxl.load_workbook(self.eligible_path)
        sheet = wb.active
        user_name = sheet.cell(row=self.last_row_number + 2, column=1).value
        phone_numbers = sheet.cell(row=self.last_row_number + 2, column=8).value
        emails = sheet.cell(row=self.last_row_number + 2, column=9).value
        to_row = sheet.cell(row=self.last_row_number + 2, column=12).value

        return user_name, phone_numbers, emails, to_row

    def companies_sheet(self, ph, emails, to_row):
        wb = openpyxl.load_workbook(self.companies_path)
        sheet = wb.active
        self.row_number = int(to_row)
        sheet.cell(row=self.row_number, column=8).value = ph
        sheet.cell(row=self.row_number, column=9).value = emails
        sheet.cell(row=self.row_number, column=12).value = self.row_number
        print(f"The provided value is save in {self.row_number}")
        wb.save(self.companies_path)

    def merger(self):
        self.enteringDetails = True
        while self.enteringDetails:
            user, phone_numbers, emails, to_row = self.eligible_sheet()
            if not user:
                print(Fore.BLACK + Back.GREEN + "No More Rows To Process")
                break
            self.companies_sheet(phone_numbers, emails, to_row)

            self.last_row_number += 1


if __name__ == "__main__":
    merger = Eligibles_Merger()
    merger.merger()
