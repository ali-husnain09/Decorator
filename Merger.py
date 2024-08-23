import openpyxl


class Details_Merger:

    def __init__(self, details_path2, companies_path2):
        self.details_path2 = details_path2
        self.companies_path2 = companies_path2
        self.last_row_number = 0
        self.r_num = 0  # Initialize r_num as a class attribute

    def companies_sheet(self):
        wb = openpyxl.load_workbook(self.companies_path2)
        sheet = wb.active
        user_name = sheet.cell(row=self.last_row_number + 2, column=1).value
        phone_numbers = sheet.cell(row=self.last_row_number + 2, column=8).value
        emails = sheet.cell(row=self.last_row_number + 2, column=9).value
        agent_name = sheet.cell(row=self.last_row_number + 2, column=10).value
        agent_address = sheet.cell(row=self.last_row_number + 2, column=11).value
        to_row = sheet.cell(row=self.last_row_number + 2, column=12).value

        return user_name, phone_numbers, emails, agent_name, agent_address, to_row

    def details_sheet(self, ph, emails, ag_N, ag_A, to_row):
        wb = openpyxl.load_workbook(self.details_path2)
        sheet = wb.active
        self.row_number = int(to_row)
        sheet.cell(row=self.row_number, column=8).value = ph
        sheet.cell(row=self.row_number, column=9).value = emails
        sheet.cell(row=self.row_number, column=10).value = ag_N
        sheet.cell(row=self.row_number, column=11).value = ag_A
        sheet.cell(row=self.row_number, column=12).value = self.row_number
        print(f"The provided value is save in {self.row_number}")
        wb.save(self.details_path2)

    def merger(self):
        self.enteringDetails = True
        while self.enteringDetails:
            user, phone_numbers, emails, agent_name, agent_address, to_row = (
                self.companies_sheet()
            )
            if not user:
                print("No More Rows To Process")
                break
            self.details_sheet(phone_numbers, emails, agent_name, agent_address, to_row)

            self.last_row_number += 1


if __name__ == "__main__":
    merger = Details_Merger()
    merger.merger()
