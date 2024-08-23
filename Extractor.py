import openpyxl
import time
import colorama
from colorama import Fore, Back

colorama.init(autoreset=True)


class NameValidations:

    def __init__(self, details_path1, companies_path1, companies_list):
        self.details_path1 = details_path1
        self.companies_path1 = companies_path1
        self.companies_list = companies_list
        self.last_row_number = 0
        self.r_num = 0  # Initialize r_num as a class attribute

    # @staticmethod
    # @njit
    def check_valid(self, name):
        name = str(name)
        valid_name = False
        valid_chars = []
        with open(self.companies_list, "r") as f:
            for line in f:
                valid_chars.append(line.strip())
        for index in valid_chars:
            if index in name.lower().split(" "):
                valid_name = True
                return valid_name
        return False

    def get_next_row_data(self):
        wb = openpyxl.load_workbook(self.details_path1)
        sheet = wb.active

        user_name = sheet.cell(row=self.last_row_number + 2, column=1).value
        property_address = sheet.cell(row=self.last_row_number + 2, column=2).value
        property_city = sheet.cell(row=self.last_row_number + 2, column=3).value
        property_state = sheet.cell(row=self.last_row_number + 2, column=4).value
        formal_address = sheet.cell(row=self.last_row_number + 2, column=5).value
        city = sheet.cell(row=self.last_row_number + 2, column=6).value
        state = sheet.cell(row=self.last_row_number + 2, column=7).value

        return (
            user_name,
            formal_address,
            city,
            state,
            property_address,
            property_city,
            property_state,
        )

    def save_value(self, sheet, value):
        sheet.cell(row=self.last_row_number + 2, column=8).value = value
        sheet.cell(row=self.last_row_number + 2, column=12).value = (
            self.last_row_number + 2
        )

    def appending_sheet(self, name, address, city, state, p_a, p_c, p_s):
        wb = openpyxl.load_workbook(self.companies_path1)
        sheet = wb.active
        sheet.cell(row=self.r_num + 2, column=1).value = name
        sheet.cell(row=self.r_num + 2, column=2).value = p_a
        sheet.cell(row=self.r_num + 2, column=3).value = p_c
        sheet.cell(row=self.r_num + 2, column=4).value = p_s
        sheet.cell(row=self.r_num + 2, column=5).value = address
        sheet.cell(row=self.r_num + 2, column=6).value = city
        sheet.cell(row=self.r_num + 2, column=7).value = state
        sheet.cell(row=self.r_num + 2, column=12).value = self.last_row_number + 2
        wb.save(self.companies_path1)
        self.r_num += 1  # Increment r_num after appending

    def company_checker(self):
        wb = openpyxl.load_workbook(self.details_path1)
        sheet = wb.active

        while True:
            user_name, formal_address, city, state, p_a, p_c, p_s = (
                self.get_next_row_data()
            )
            if not user_name:
                print("No More Rows To Process")
                print(
                    f"{Fore.BLACK + Back.LIGHTGREEN_EX}Total Rows are: {self.last_row_number}"
                )
                break

            if self.check_valid(user_name):
                print(f"Company Found: {user_name} is valid.")
                self.save_value(sheet, "It's a Company")
                self.appending_sheet(
                    user_name, formal_address, city, state, p_a, p_c, p_s
                )

            self.last_row_number += 1

        wb.save(self.details_path1)


details_path1 = "Extractor/details.xlsx"
companies_path1 = "Extractor/companies.xlsx"
companies_list = "Extractor/companies.txt"

if __name__ == "__main__":
    # Start the timer
    start_time = time.time()
    validation_obj = NameValidations(details_path1, companies_path1, companies_list)
    validation_obj.company_checker()
    # End the timer
    end_time = time.time()

    # Calculate the execution time
    execution_time = end_time - start_time

    print(f"\nExecution time:{Fore.LIGHTYELLOW_EX} {execution_time} seconds")
